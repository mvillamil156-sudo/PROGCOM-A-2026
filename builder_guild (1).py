"""
                                                                              
Instalación:  pip install openpyxl pillow                                  

"""

import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import tkinter.font as tkfont
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os, math, random, datetime
from pathlib import Path

try:
    from PIL import Image, ImageDraw, ImageFont, ImageTk
    PIL_OK = True
except ImportError:
    PIL_OK = False

# ─── RUTA DEL EXCEL ────────────────────────────────────────────────
EXCEL_PATH = Path(__file__).parent / "constructores.xlsx"

# ─── ROLES Y CONFIGURACIÓN ─────────────────────────────────────────
ROLES = {
    "Arquitecto":   {"color": "#f59e0b", "emoji": "🏛️",  "desc": "Diseña estructuras épicas"},
    "Minero":       {"color": "#94a3b8", "emoji": "⛏️",  "desc": "Extrae recursos valiosos"},
    "Herrero":      {"color": "#ef4444", "emoji": "🔨",  "desc": "Forja herramientas y armas"},
    "Agricultor":   {"color": "#22c55e", "emoji": "🌾",  "desc": "Cultiva y provee alimentos"},
    "Explorador":   {"color": "#06b6d4", "emoji": "🧭",  "desc": "Descubre nuevos territorios"},
    "Mago":         {"color": "#8b5cf6", "emoji": "🔮",  "desc": "Domina las artes arcanas"},
    "Guerrero":     {"color": "#dc2626", "emoji": "⚔️",  "desc": "Protege la aldea"},
    "Comerciante":  {"color": "#f97316", "emoji": "💰",  "desc": "Gestiona el comercio"},
    "Constructor":  {"color": "#3b82f6", "emoji": "🧱",  "desc": "Erige edificios y murallas"},
    "Alquimista":   {"color": "#10b981", "emoji": "⚗️",  "desc": "Crea pociones y elíxires"},
}

# Niveles basados en XP
NIVELES = [
    (0,     "Novato",      "⬜"),
    (100,   "Aprendiz",    "🟩"),
    (300,   "Oficial",     "🟦"),
    (600,   "Experto",     "🟨"),
    (1000,  "Maestro",     "🟧"),
    (1500,  "Gran Maestro","🟥"),
    (2500,  "Leyenda",     "🌟"),
    (5000,  "Dios",        "💎"),
]

QUESTS_DISPONIBLES = [
    "Primera Construcción",
    "Recolectar 100 recursos",
    "Completar un edificio épico",
    "Defender la aldea",
    "Explorar 5 regiones",
    "Forjar el hacha legendaria",
    "Cultivar el jardín secreto",
    "Comerciar con 3 reinos",
    "Descifrar el pergamino antiguo",
    "Construir el castillo",
    "Minar el cristal arcano",
    "Entrenar 10 guerreros",
    "Completar la torre del mago",
    "Fundar una colonia",
    "Escalar la montaña de fuego",
]

KINKET_ITEMS = [
    "Pico de Diamante",
    "Espada Legendaria",
    "Escudo Dorado",
    "Poción de Poder",
    "Mapa del Tesoro",
    "Hacha Encantada",
    "Armadura de Dragón",
    "Libro de Hechizos",
    "Amuleto del Constructor",
    "Corona del Gremio",
    "Martillo de Thor",
    "Llave de la Mazmorra",
    "Cristal de Visión",
    "Botas de Velocidad",
    "Anillo de Protección",
]

# ─── PALETA ────────────────────────────────────────────────────────
BG      = "#0f172a"
BG2     = "#1e293b"
BG3     = "#334155"
BORDER  = "#475569"
GOLD    = "#f59e0b"
GOLD2   = "#fcd34d"
CYAN    = "#06b6d4"
GREEN   = "#22c55e"
RED     = "#ef4444"
PURPLE  = "#8b5cf6"
WHITE   = "#f1f5f9"
GRAY    = "#94a3b8"
DIM     = "#475569"

# ══════════════════════════════════════════════════════════════════
#  EXCEL — Capa de datos
# ══════════════════════════════════════════════════════════════════

class ExcelDB:
    """Gestiona lectura/escritura del Excel de constructores."""

    COLS = ["ID", "Nombre", "Rol", "EXP", "Kinket", "Quests", "Fecha_Ingreso", "Notas"]

    def __init__(self, ruta: Path):
        self.ruta = ruta
        if not ruta.exists():
            self._crear_excel_inicial()

    # ── Crear Excel con datos demo ──────────────────────────────
    def _crear_excel_inicial(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Constructores"

        # Encabezados estilizados
        headers = self.COLS
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
            cell.fill      = PatternFill("solid", start_color="1E293B")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[1].height = 28

        # Anchos de columna
        widths = [8, 20, 15, 10, 15, 50, 14, 30]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = w

        # Datos de ejemplo
        demo = [
            [1, "Carlos Ruiz",    "Arquitecto",  850,  "Pico de Diamante,Mapa del Tesoro",
             "Primera Construcción,Construir el castillo,Explorar 5 regiones",
             "2024-01-15", "Especialista en puentes"],
            [2, "María López",    "Mago",         1250, "Libro de Hechizos,Cristal de Visión,Amuleto del Constructor",
             "Primera Construcción,Completar la torre del mago,Descifrar el pergamino antiguo",
             "2024-02-01", "Maestra de pociones"],
            [3, "Pedro Gómez",    "Guerrero",     320,  "Espada Legendaria,Escudo Dorado",
             "Primera Construcción,Defender la aldea",
             "2024-03-10", "Protector del norte"],
            [4, "Ana Martínez",   "Comerciante",  670,  "Mapa del Tesoro,Poción de Poder",
             "Primera Construcción,Comerciar con 3 reinos,Recolectar 100 recursos",
             "2024-03-22", "Red de comercio en el este"],
            [5, "Luis Herrera",   "Minero",       150,  "Pico de Diamante",
             "Primera Construcción",
             "2024-04-05", "Nuevo en el gremio"],
            [6, "Sofia Castro",   "Alquimista",   1800, "Libro de Hechizos,Cristal de Visión,Anillo de Protección,Amuleto del Constructor",
             "Primera Construcción,Completar la torre del mago,Forjar el hacha legendaria,Minar el cristal arcano",
             "2023-11-20", "Gran Maestra de alquimia"],
            [7, "Jorge Díaz",     "Constructor",  540,  "Martillo de Thor,Armadura de Dragón",
             "Primera Construcción,Completar un edificio épico,Construir el castillo",
             "2024-01-30", "Especialista en murallas"],
            [8, "Laura Torres",   "Exploradora",  980,  "Mapa del Tesoro,Botas de Velocidad,Llave de la Mazmorra",
             "Primera Construcción,Explorar 5 regiones,Fundar una colonia,Escalar la montaña de fuego",
             "2023-12-01", "Exploradora del sur"],
        ]

        for row_data in demo:
            ws.append(row_data)

        # Hoja de Quests
        wq = wb.create_sheet("Quests_Log")
        wq.append(["ID_Constructor", "Nombre_Constructor", "Quest", "Fecha_Completada", "EXP_Ganada"])
        wq.column_dimensions["A"].width = 16
        wq.column_dimensions["B"].width = 22
        wq.column_dimensions["C"].width = 35
        wq.column_dimensions["D"].width = 18
        wq.column_dimensions["E"].width = 14

        # Hoja Kinket_Log
        wk = wb.create_sheet("Kinket_Log")
        wk.append(["ID_Constructor", "Nombre_Constructor", "Item", "Fecha_Obtenido", "Tipo"])

        wb.save(self.ruta)

    # ── Leer todos los constructores ───────────────────────────
    def leer_todos(self) -> list[dict]:
        wb = load_workbook(self.ruta, data_only=True)
        ws = wb["Constructores"]
        constructores = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            quests   = [q.strip() for q in str(row[5]).split(",") if q.strip() and q.strip() != "None"]
            kinket   = [k.strip() for k in str(row[4]).split(",") if k.strip() and k.strip() != "None"]
            constructores.append({
                "id":           int(row[0]) if row[0] else 0,
                "nombre":       str(row[1]) if row[1] else "",
                "rol":          str(row[2]) if row[2] else "Constructor",
                "exp":          int(row[3]) if row[3] else 0,
                "kinket":       kinket,
                "quests":       quests,
                "fecha_ingreso":str(row[6]) if row[6] else "",
                "notas":        str(row[7]) if row[7] else "",
            })
        return constructores

    def leer_uno(self, cid: int) -> dict | None:
        for c in self.leer_todos():
            if c["id"] == cid:
                return c
        return None

    # ── Guardar / actualizar constructor ──────────────────────
    def guardar(self, datos: dict):
        wb = load_workbook(self.ruta)
        ws = wb["Constructores"]

        # Buscar fila existente
        fila_existente = None
        for row in ws.iter_rows(min_row=2):
            if row[0].value == datos["id"]:
                fila_existente = row[0].row
                break

        quests_str = ",".join(datos.get("quests", []))
        kinket_str = ",".join(datos.get("kinket", []))
        valores = [
            datos["id"], datos["nombre"], datos["rol"],
            datos["exp"], kinket_str, quests_str,
            datos.get("fecha_ingreso", str(datetime.date.today())),
            datos.get("notas", ""),
        ]

        if fila_existente:
            for i, v in enumerate(valores, 1):
                ws.cell(row=fila_existente, column=i, value=v)
        else:
            ws.append(valores)

        wb.save(self.ruta)

    # ── Eliminar constructor ───────────────────────────────────
    def eliminar(self, cid: int):
        wb = load_workbook(self.ruta)
        ws = wb["Constructores"]
        for row in ws.iter_rows(min_row=2):
            if row[0].value == cid:
                ws.delete_rows(row[0].row)
                break
        wb.save(self.ruta)

    # ── Siguiente ID ───────────────────────────────────────────
    def siguiente_id(self) -> int:
        todos = self.leer_todos()
        return max((c["id"] for c in todos), default=0) + 1

    # ── Log de quest ──────────────────────────────────────────
    def log_quest(self, constructor: dict, quest: str, exp_ganada: int):
        wb = load_workbook(self.ruta)
        ws = wb["Quests_Log"]
        ws.append([
            constructor["id"], constructor["nombre"], quest,
            str(datetime.date.today()), exp_ganada,
        ])
        wb.save(self.ruta)

    # ── Log de kinket ─────────────────────────────────────────
    def log_kinket(self, constructor: dict, item: str):
        wb = load_workbook(self.ruta)
        ws = wb["Kinket_Log"]
        ws.append([
            constructor["id"], constructor["nombre"], item,
            str(datetime.date.today()), "Obtenido",
        ])
        wb.save(self.ruta)


# ══════════════════════════════════════════════════════════════════
#  LÓGICA DE NIVELES
# ══════════════════════════════════════════════════════════════════

def get_nivel(exp: int) -> tuple[int, str, str, int, int]:
    """Retorna (num_nivel, nombre, emoji, exp_actual_en_nivel, exp_para_siguiente)"""
    nivel_idx = 0
    for i, (xp_req, _, _) in enumerate(NIVELES):
        if exp >= xp_req:
            nivel_idx = i
    num   = nivel_idx + 1
    nombre = NIVELES[nivel_idx][1]
    emoji  = NIVELES[nivel_idx][2]
    exp_base = NIVELES[nivel_idx][0]
    exp_sig  = NIVELES[nivel_idx + 1][0] if nivel_idx + 1 < len(NIVELES) else NIVELES[-1][0]
    return num, nombre, emoji, exp - exp_base, exp_sig - exp_base


# ══════════════════════════════════════════════════════════════════
#  GENERADOR DE AVATARES  (PIL)
# ══════════════════════════════════════════════════════════════════

class AvatarGen:
    """Genera avatares procedurales basados en rol y nombre."""

    SIZE = 120

    @staticmethod
    def generar(rol: str, nombre: str, size: int = 120) -> "ImageTk.PhotoImage | None":
        if not PIL_OK:
            return None

        cfg  = ROLES.get(rol, ROLES["Constructor"])
        col  = cfg["color"]

        # Color base del rol
        h = col.lstrip("#")
        rc, gc, bc = int(h[0:2],16), int(h[2:4],16), int(h[4:6],16)

        img  = Image.new("RGBA", (size, size), (0,0,0,0))
        draw = ImageDraw.Draw(img)

        # Fondo circular con gradiente simulado
        for r2 in range(size//2, 0, -1):
            t   = r2 / (size//2)
            r_c = int(rc * t + 10*(1-t))
            g_c = int(gc * t + 8 *(1-t))
            b_c = int(bc * t + 20*(1-t))
            cx  = size//2
            draw.ellipse([cx-r2, cx-r2, cx+r2, cx+r2],
                         fill=(r_c, g_c, b_c, 255))

        # Anillo exterior
        ring_color = (
            min(255, rc + 80),
            min(255, gc + 80),
            min(255, bc + 80), 255)
        draw.ellipse([3, 3, size-4, size-4],
                     outline=ring_color, width=3)

        # Inicial del nombre grande
        inicial = nombre[0].upper() if nombre else "?"
        try:
            fnt = ImageFont.truetype("arial.ttf", size=size//2)
        except Exception:
            fnt = ImageFont.load_default()

        # Sombra
        bb   = draw.textbbox((0,0), inicial, font=fnt)
        tw   = bb[2] - bb[0]
        th   = bb[3] - bb[1]
        tx   = (size - tw) // 2 - bb[0]
        ty   = (size - th) // 2 - bb[1] - 4
        draw.text((tx+2, ty+2), inicial, font=fnt, fill=(0,0,0,120))
        draw.text((tx,   ty),   inicial, font=fnt, fill=(255,255,255,240))

        # Insignia de rol (esquina inferior derecha)
        badge_r = size // 6
        bx = size - badge_r - 4
        by = size - badge_r - 4
        draw.ellipse([bx-badge_r, by-badge_r, bx+badge_r, by+badge_r],
                     fill=(15, 23, 42, 240),
                     outline=ring_color, width=2)

        return ImageTk.PhotoImage(img)

    @staticmethod
    def generar_mini(rol: str, nombre: str, size: int = 44) -> "ImageTk.PhotoImage | None":
        return AvatarGen.generar(rol, nombre, size)


# ══════════════════════════════════════════════════════════════════
#  WIDGETS REUTILIZABLES
# ══════════════════════════════════════════════════════════════════

def mk_btn(parent, text, cmd, bg=GOLD, fg=BG, font=None, **kw):
    f = font or ("Helvetica", 9, "bold")
    b = tk.Button(parent, text=text, command=cmd,
                  bg=bg, fg=fg, relief="flat",
                  font=f, padx=12, pady=6,
                  activebackground=bg, activeforeground=fg,
                  cursor="hand2", **kw)
    def _h(c=bg): return f"#{min(255,int(c.lstrip('#')[0:2],16)+40):02x}" \
                         f"{min(255,int(c.lstrip('#')[2:4],16)+40):02x}" \
                         f"{min(255,int(c.lstrip('#')[4:6],16)+40):02x}"
    b.bind("<Enter>", lambda e: b.config(bg=_h()))
    b.bind("<Leave>", lambda e: b.config(bg=bg))
    return b

def sep(parent, color=BORDER, pady=4):
    f = tk.Frame(parent, bg=color, height=1)
    f.pack(fill=tk.X, pady=pady)
    return f

def lbl(parent, text, fg=WHITE, font=None, **kw):
    f = font or ("Helvetica", 10)
    return tk.Label(parent, text=text, bg=kw.pop("bg", BG2),
                    fg=fg, font=f, **kw)


# ══════════════════════════════════════════════════════════════════
#  BARRA DE XP ANIMADA
# ══════════════════════════════════════════════════════════════════

class XPBar(tk.Canvas):
    def __init__(self, parent, width=300, height=16, **kw):
        super().__init__(parent, width=width, height=height,
                         bg=BG3, highlightthickness=1,
                         highlightbackground=BORDER, **kw)
        self._width = width
        self._height = height
        self._val = 0.0
        self._target = 0.0
        self._color = GOLD
        self._draw()

    def set_value(self, porcentaje: float, color: str = GOLD, animate: bool = True):
        self._target = max(0.0, min(1.0, porcentaje))
        self._color  = color
        if animate:
            self._animar()
        else:
            self._val = self._target
            self._draw()

    def _animar(self):
        diff = self._target - self._val
        if abs(diff) > 0.005:
            self._val += diff * 0.15
            self._draw()
            self.after(16, self._animar)
        else:
            self._val = self._target
            self._draw()

    def _draw(self):
        self.delete("all")
        self.create_rectangle(0, 0, self._width, self._height,
                               fill=BG3, outline="")
        fill_w = max(4, int(self._width * self._val))
        # Gradiente simulado con múltiples rectángulos
        h = self._color.lstrip("#")
        r0,g0,b0 = int(h[0:2],16), int(h[2:4],16), int(h[4:6],16)
        segments = 20
        for i in range(segments):
            t   = i / segments
            seg_x0 = int(fill_w * i / segments)
            seg_x1 = int(fill_w * (i+1) / segments)
            brightness = 0.7 + 0.3 * (1 - abs(t - 0.6))
            rc = min(255, int(r0 * brightness))
            gc = min(255, int(g0 * brightness))
            bc = min(255, int(b0 * brightness))
            self.create_rectangle(seg_x0, 2, seg_x1, self._height-2,
                                   fill=f"#{rc:02x}{gc:02x}{bc:02x}", outline="")
        # Brillo superior
        self.create_rectangle(2, 2, fill_w-2, self._height//2,
                               fill="", outline="",
                               stipple="gray25")


# ══════════════════════════════════════════════════════════════════
#  VENTANA DETALLE DEL CONSTRUCTOR
# ══════════════════════════════════════════════════════════════════

class DetalleWindow(tk.Toplevel):
    def __init__(self, master, db: ExcelDB, cid: int, on_close=None):
        super().__init__(master)
        self.db       = db
        self.cid      = cid
        self.on_close = on_close
        self.configure(bg=BG)
        self.resizable(True, True)
        self.geometry("780x680")
        self.grab_set()
        self.transient(master)
        self._avatar_ref = None
        self._cargar_y_construir()

    def _cargar_y_construir(self):
        self.datos = self.db.leer_uno(self.cid)
        if not self.datos:
            self.destroy(); return
        self.title(f"🏰  {self.datos['nombre']}  —  Builder Guild")
        self._build()

    def _build(self):
        for w in self.winfo_children(): w.destroy()
        d    = self.datos
        rol  = d["rol"]
        cfg  = ROLES.get(rol, ROLES["Constructor"])
        col  = cfg["color"]
        exp  = d["exp"]
        num_nv, nv_nombre, nv_emoji, exp_en_nv, exp_para_nv = get_nivel(exp)

        # ── HEADER ──────────────────────────────────────────────
        hdr = tk.Frame(self, bg=BG2, padx=20, pady=14)
        hdr.pack(fill=tk.X)

        # Avatar
        if PIL_OK:
            self._avatar_ref = AvatarGen.generar(rol, d["nombre"], 90)
            if self._avatar_ref:
                tk.Label(hdr, image=self._avatar_ref, bg=BG2).pack(side=tk.LEFT, padx=(0,16))

        # Info principal
        info_f = tk.Frame(hdr, bg=BG2)
        info_f.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        tk.Label(info_f, text=d["nombre"],
                 bg=BG2, fg=WHITE,
                 font=("Georgia", 20, "bold")).pack(anchor="w")

        rol_f = tk.Frame(info_f, bg=BG2)
        rol_f.pack(anchor="w", pady=2)
        tk.Label(rol_f, text=f"{cfg['emoji']}  {rol}",
                 bg=col, fg=BG,
                 font=("Helvetica", 10, "bold"),
                 padx=10, pady=3).pack(side=tk.LEFT)
        tk.Label(rol_f, text=f"   {cfg['desc']}",
                 bg=BG2, fg=GRAY,
                 font=("Helvetica", 9, "italic")).pack(side=tk.LEFT)

        # Nivel badge
        nv_f = tk.Frame(info_f, bg=BG2)
        nv_f.pack(anchor="w", pady=2)
        tk.Label(nv_f, text=f"{nv_emoji} Nivel {num_nv} — {nv_nombre}",
                 bg=BG2, fg=col,
                 font=("Helvetica", 11, "bold")).pack(side=tk.LEFT)
        tk.Label(nv_f, text=f"   📅 Ingreso: {d.get('fecha_ingreso','')}",
                 bg=BG2, fg=GRAY,
                 font=("Helvetica", 9)).pack(side=tk.LEFT, padx=20)

        # Acciones
        act_f = tk.Frame(hdr, bg=BG2)
        act_f.pack(side=tk.RIGHT, anchor="ne")
        mk_btn(act_f, "✏️ Editar",  self._editar, BG3, WHITE).pack(pady=2, fill=tk.X)
        mk_btn(act_f, "🗑️ Eliminar",self._eliminar, RED, WHITE).pack(pady=2, fill=tk.X)

        tk.Frame(self, bg=BORDER, height=1).pack(fill=tk.X)

        # ── CUERPO PRINCIPAL ────────────────────────────────────
        body = tk.Frame(self, bg=BG)
        body.pack(fill=tk.BOTH, expand=True, padx=16, pady=12)

        left_col = tk.Frame(body, bg=BG)
        left_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0,8))

        right_col = tk.Frame(body, bg=BG, width=280)
        right_col.pack(side=tk.RIGHT, fill=tk.Y)
        right_col.pack_propagate(False)

        # ── STATS ──────────────────────────────────────────────
        self._build_stats(left_col, d, exp, num_nv, nv_nombre, col, exp_en_nv, exp_para_nv)

        # ── QUESTS ─────────────────────────────────────────────
        self._build_quests(left_col, d)

        # ── KINKET ─────────────────────────────────────────────
        self._build_kinket(right_col, d)

    def _build_stats(self, parent, d, exp, num_nv, nv_nombre, col, exp_en_nv, exp_para_nv):
        sf = tk.Frame(parent, bg=BG2, padx=14, pady=12)
        sf.pack(fill=tk.X, pady=(0,8))

        tk.Label(sf, text="📊  ESTADÍSTICAS",
                 bg=BG2, fg=GOLD, font=("Helvetica",10,"bold")).pack(anchor="w")
        tk.Frame(sf, bg=GOLD, height=1).pack(fill=tk.X, pady=(3,8))

        # Grid de stats
        grid_f = tk.Frame(sf, bg=BG2)
        grid_f.pack(fill=tk.X)

        stats_data = [
            ("⭐ EXP Total",     f"{exp:,}", GOLD),
            ("🏅 Nivel",          f"{num_nv} — {nv_nombre}", col),
            ("🎯 Quests",         str(len(d["quests"])), CYAN),
            ("🎒 Items Kinket",   str(len(d["kinket"])), PURPLE),
        ]

        for i, (label, valor, c) in enumerate(stats_data):
            r, col_i = divmod(i, 2)
            cell = tk.Frame(grid_f, bg=BG3, padx=12, pady=8)
            cell.grid(row=r, column=col_i, padx=4, pady=4, sticky="ew")
            grid_f.columnconfigure(col_i, weight=1)
            tk.Label(cell, text=label, bg=BG3, fg=GRAY,
                     font=("Helvetica",8)).pack(anchor="w")
            tk.Label(cell, text=valor, bg=BG3, fg=c,
                     font=("Courier New",16,"bold")).pack(anchor="w")

        # Barra de XP
        xp_f = tk.Frame(sf, bg=BG2)
        xp_f.pack(fill=tk.X, pady=(10,0))
        pct = exp_en_nv / exp_para_nv if exp_para_nv > 0 else 1.0

        sig_idx  = min(len(NIVELES)-1, sum(1 for x,_,_ in NIVELES if exp >= x))
        sig_nom  = NIVELES[sig_idx][1] if sig_idx < len(NIVELES) else "Máximo"

        tk.Label(xp_f, text=f"Progreso al siguiente nivel ({sig_nom})",
                 bg=BG2, fg=GRAY, font=("Helvetica",8)).pack(anchor="w")

        bar_row = tk.Frame(xp_f, bg=BG2)
        bar_row.pack(fill=tk.X, pady=2)
        xp_bar = XPBar(bar_row, width=360, height=14)
        xp_bar.pack(side=tk.LEFT)
        tk.Label(bar_row,
                 text=f"  {exp_en_nv:,} / {exp_para_nv:,} XP",
                 bg=BG2, fg=col, font=("Courier New",9)).pack(side=tk.LEFT)
        xp_bar.set_value(pct, col)

        # Notas
        if d.get("notas"):
            tk.Label(sf, text=f"📝  {d['notas']}",
                     bg=BG2, fg=GRAY,
                     font=("Helvetica",9,"italic")).pack(anchor="w", pady=(8,0))

    def _build_quests(self, parent, d):
        qf = tk.Frame(parent, bg=BG2, padx=14, pady=12)
        qf.pack(fill=tk.BOTH, expand=True, pady=(0,8))

        hdr = tk.Frame(qf, bg=BG2)
        hdr.pack(fill=tk.X)
        tk.Label(hdr, text=f"🎯  QUESTS COMPLETADAS  ({len(d['quests'])})",
                 bg=BG2, fg=CYAN, font=("Helvetica",10,"bold")).pack(side=tk.LEFT)
        mk_btn(hdr, "+ Agregar Quest",
               lambda: self._agregar_quest(), CYAN, BG,
               font=("Helvetica",8,"bold")).pack(side=tk.RIGHT)

        tk.Frame(qf, bg=CYAN, height=1).pack(fill=tk.X, pady=(3,6))

        if not d["quests"]:
            tk.Label(qf, text="Sin quests completadas aún",
                     bg=BG2, fg=DIM, font=("Helvetica",9,"italic")).pack(pady=8)
        else:
            wrap = tk.Frame(qf, bg=BG2)
            wrap.pack(fill=tk.BOTH, expand=True)
            canvas = tk.Canvas(wrap, bg=BG2, highlightthickness=0, height=140)
            vsb    = ttk.Scrollbar(wrap, orient="vertical", command=canvas.yview)
            canvas.configure(yscrollcommand=vsb.set)
            vsb.pack(side=tk.RIGHT, fill=tk.Y)
            canvas.pack(fill=tk.BOTH, expand=True)

            inner = tk.Frame(canvas, bg=BG2)
            canvas.create_window((0,0), window=inner, anchor="nw")
            inner.bind("<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

            cols_q = 2
            for i, quest in enumerate(d["quests"]):
                r, c = divmod(i, cols_q)
                qcard = tk.Frame(inner, bg=BG3, padx=8, pady=4)
                qcard.grid(row=r, column=c, padx=3, pady=2, sticky="ew")
                inner.columnconfigure(c, weight=1)
                tk.Label(qcard, text=f"✅  {quest}",
                         bg=BG3, fg=WHITE,
                         font=("Helvetica",9), wraplength=200,
                         justify="left").pack(anchor="w")

    def _build_kinket(self, parent, d):
        kf = tk.Frame(parent, bg=BG2, padx=12, pady=12)
        kf.pack(fill=tk.BOTH, expand=True)

        hdr = tk.Frame(kf, bg=BG2)
        hdr.pack(fill=tk.X)
        tk.Label(hdr, text=f"🎒  KINKET  ({len(d['kinket'])})",
                 bg=BG2, fg=PURPLE, font=("Helvetica",10,"bold")).pack(side=tk.LEFT)
        mk_btn(hdr, "+ Item",
               lambda: self._agregar_kinket(), PURPLE, WHITE,
               font=("Helvetica",8,"bold")).pack(side=tk.RIGHT)

        tk.Frame(kf, bg=PURPLE, height=1).pack(fill=tk.X, pady=(3,6))

        if not d["kinket"]:
            tk.Label(kf, text="Kinket vacío",
                     bg=BG2, fg=DIM,
                     font=("Helvetica",9,"italic")).pack(pady=8)
        else:
            for item in d["kinket"]:
                row = tk.Frame(kf, bg=BG3, padx=8, pady=5)
                row.pack(fill=tk.X, pady=2)
                tk.Label(row, text=f"🔹 {item}",
                         bg=BG3, fg=WHITE,
                         font=("Helvetica",9)).pack(side=tk.LEFT, fill=tk.X, expand=True)
                tk.Button(row, text="✕",
                          command=lambda i=item: self._quitar_kinket(i),
                          bg=BG3, fg=RED, relief="flat",
                          font=("Helvetica",8), cursor="hand2",
                          activebackground=BG3, activeforeground=RED).pack(side=tk.RIGHT)

    # ── ACCIONES ───────────────────────────────────────────────
    def _agregar_quest(self):
        pendientes = [q for q in QUESTS_DISPONIBLES
                      if q not in self.datos["quests"]]
        if not pendientes:
            messagebox.showinfo("Sin quests", "¡Todas las quests completadas!")
            return
        QuestSelector(self, pendientes, self._on_quest_seleccionada)

    def _on_quest_seleccionada(self, quest: str, exp_bonus: int):
        self.datos["quests"].append(quest)
        self.datos["exp"] += exp_bonus
        self.db.guardar(self.datos)
        self.db.log_quest(self.datos, quest, exp_bonus)
        messagebox.showinfo("¡Quest completada!",
                            f"✅ {quest}\n+{exp_bonus} EXP ganados!")
        self._cargar_y_construir()
        if self.on_close:
            self.on_close()

    def _agregar_kinket(self):
        disponibles = [k for k in KINKET_ITEMS
                       if k not in self.datos["kinket"]]
        if not disponibles:
            messagebox.showinfo("Kinket lleno", "Ya tienes todos los items disponibles.")
            return
        KinketSelector(self, disponibles, self._on_kinket_seleccionado)

    def _on_kinket_seleccionado(self, item: str):
        self.datos["kinket"].append(item)
        self.db.guardar(self.datos)
        self.db.log_kinket(self.datos, item)
        messagebox.showinfo("¡Item obtenido!", f"🎒 {item} añadido al kinket!")
        self._cargar_y_construir()
        if self.on_close:
            self.on_close()

    def _quitar_kinket(self, item: str):
        if messagebox.askyesno("Quitar item", f"¿Quitar '{item}' del kinket?"):
            self.datos["kinket"].remove(item)
            self.db.guardar(self.datos)
            self._cargar_y_construir()
            if self.on_close: self.on_close()

    def _editar(self):
        EditarWindow(self, self.db, self.datos,
                     on_save=lambda: (self._cargar_y_construir(),
                                      self.on_close and self.on_close()))

    def _eliminar(self):
        if messagebox.askyesno("Eliminar constructor",
                               f"¿Eliminar a {self.datos['nombre']} definitivamente?"):
            self.db.eliminar(self.cid)
            if self.on_close: self.on_close()
            self.destroy()


# ══════════════════════════════════════════════════════════════════
#  SELECTOR DE QUEST
# ══════════════════════════════════════════════════════════════════

class QuestSelector(tk.Toplevel):
    EXP_POR_QUEST = 75

    def __init__(self, master, quests: list[str], callback):
        super().__init__(master)
        self.callback = callback
        self.title("🎯 Seleccionar Quest")
        self.configure(bg=BG)
        self.geometry("400x480")
        self.grab_set()
        self.resizable(False, True)

        tk.Label(self, text="🎯  Elige una quest completada",
                 bg=BG, fg=CYAN,
                 font=("Georgia", 13, "bold")).pack(pady=(16,8))
        tk.Frame(self, bg=BORDER, height=1).pack(fill=tk.X, padx=20)

        wrap = tk.Frame(self, bg=BG)
        wrap.pack(fill=tk.BOTH, expand=True, padx=16, pady=8)

        canvas = tk.Canvas(wrap, bg=BG, highlightthickness=0)
        vsb    = ttk.Scrollbar(wrap, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(fill=tk.BOTH, expand=True)

        inner = tk.Frame(canvas, bg=BG)
        canvas.create_window((0,0), window=inner, anchor="nw")
        inner.bind("<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        for quest in quests:
            exp_r = random.randint(50, 150)
            row   = tk.Frame(inner, bg=BG2, padx=10, pady=8, cursor="hand2")
            row.pack(fill=tk.X, pady=2)
            tk.Label(row, text=f"🎯  {quest}",
                     bg=BG2, fg=WHITE,
                     font=("Helvetica",10), anchor="w").pack(side=tk.LEFT, fill=tk.X, expand=True)
            tk.Label(row, text=f"+{exp_r} XP",
                     bg=BG2, fg=GOLD,
                     font=("Courier New",9,"bold")).pack(side=tk.RIGHT)
            row.bind("<Button-1>", lambda e, q=quest, x=exp_r: self._seleccionar(q, x))
            for w in row.winfo_children():
                w.bind("<Button-1>", lambda e, q=quest, x=exp_r: self._seleccionar(q, x))
            row.bind("<Enter>", lambda e, f=row: f.config(bg=BG3) or [w.config(bg=BG3) for w in f.winfo_children()])
            row.bind("<Leave>", lambda e, f=row: f.config(bg=BG2) or [w.config(bg=BG2) for w in f.winfo_children()])

        mk_btn(self, "Cancelar", self.destroy, BG3, GRAY).pack(pady=8)

    def _seleccionar(self, quest, exp):
        self.destroy()
        self.callback(quest, exp)


# ══════════════════════════════════════════════════════════════════
#  SELECTOR DE KINKET
# ══════════════════════════════════════════════════════════════════

class KinketSelector(tk.Toplevel):
    def __init__(self, master, items: list[str], callback):
        super().__init__(master)
        self.callback = callback
        self.title("🎒 Agregar al Kinket")
        self.configure(bg=BG)
        self.geometry("360x400")
        self.grab_set()
        self.resizable(False, True)

        tk.Label(self, text="🎒  Elige un item para tu kinket",
                 bg=BG, fg=PURPLE,
                 font=("Georgia",12,"bold")).pack(pady=(14,6))
        tk.Frame(self, bg=BORDER, height=1).pack(fill=tk.X, padx=20)

        wrap = tk.Frame(self, bg=BG)
        wrap.pack(fill=tk.BOTH, expand=True, padx=14, pady=6)

        for item in items:
            row = tk.Frame(wrap, bg=BG2, padx=10, pady=7, cursor="hand2")
            row.pack(fill=tk.X, pady=2)
            tk.Label(row, text=f"🔹  {item}",
                     bg=BG2, fg=WHITE, font=("Helvetica",10)).pack(anchor="w")
            row.bind("<Button-1>", lambda e, i=item: self._sel(i))
            row.winfo_children()[0].bind("<Button-1>", lambda e, i=item: self._sel(i))
            row.bind("<Enter>", lambda e, f=row: f.config(bg=BG3) or [w.config(bg=BG3) for w in f.winfo_children()])
            row.bind("<Leave>", lambda e, f=row: f.config(bg=BG2) or [w.config(bg=BG2) for w in f.winfo_children()])

        mk_btn(self, "Cancelar", self.destroy, BG3, GRAY).pack(pady=8)

    def _sel(self, item):
        self.destroy()
        self.callback(item)


# ══════════════════════════════════════════════════════════════════
#  FORMULARIO CREAR / EDITAR CONSTRUCTOR
# ══════════════════════════════════════════════════════════════════

class EditarWindow(tk.Toplevel):
    def __init__(self, master, db: ExcelDB, datos: dict | None = None, on_save=None):
        super().__init__(master)
        self.db      = db
        self.datos   = datos
        self.on_save = on_save
        self.modo    = "editar" if datos else "crear"
        self.title("✏️ Editar Constructor" if self.modo == "editar" else "➕ Nuevo Constructor")
        self.configure(bg=BG)
        self.geometry("480x560")
        self.resizable(False, True)
        self.grab_set()
        self._build()

    def _build(self):
        d = self.datos or {}

        tk.Label(self,
                 text="✏️  Editar Constructor" if self.modo=="editar" else "➕  Nuevo Constructor",
                 bg=BG, fg=GOLD,
                 font=("Georgia",14,"bold")).pack(pady=(16,8))
        tk.Frame(self, bg=BORDER, height=1).pack(fill=tk.X, padx=20)

        form = tk.Frame(self, bg=BG, padx=24)
        form.pack(fill=tk.BOTH, expand=True, pady=12)

        def row_entry(label, default="", fg=CYAN):
            f = tk.Frame(form, bg=BG)
            f.pack(fill=tk.X, pady=5)
            tk.Label(f, text=label, bg=BG, fg=GRAY,
                     font=("Helvetica",9), width=14, anchor="w").pack(side=tk.LEFT)
            var = tk.StringVar(value=str(default))
            e = tk.Entry(f, textvariable=var, bg=BG3, fg=fg,
                         insertbackground=fg, relief="solid",
                         font=("Helvetica",10), bd=1, width=26)
            e.pack(side=tk.LEFT, ipady=5)
            return var

        self._v_nombre = row_entry("Nombre:",      d.get("nombre",""))
        self._v_exp    = row_entry("EXP:",         d.get("exp", 0), GOLD)
        self._v_notas  = row_entry("Notas:",       d.get("notas",""), GRAY)

        # Rol selector
        rol_f = tk.Frame(form, bg=BG)
        rol_f.pack(fill=tk.X, pady=5)
        tk.Label(rol_f, text="Rol:", bg=BG, fg=GRAY,
                 font=("Helvetica",9), width=14, anchor="w").pack(side=tk.LEFT)
        self._v_rol = tk.StringVar(value=d.get("rol","Constructor"))
        rol_cb = ttk.Combobox(rol_f, textvariable=self._v_rol,
                               values=list(ROLES.keys()),
                               state="readonly", width=24,
                               font=("Helvetica",10))
        rol_cb.pack(side=tk.LEFT)

        # Vista previa del rol
        prev_f = tk.Frame(form, bg=BG2, padx=10, pady=8)
        prev_f.pack(fill=tk.X, pady=6)
        self._prev_lbl = tk.Label(prev_f, text="", bg=BG2,
                                   font=("Helvetica",10))
        self._prev_lbl.pack()
        self._avatar_prev = None
        self._prev_avatar_lbl = tk.Label(prev_f, bg=BG2)
        self._prev_avatar_lbl.pack()

        def update_preview(*_):
            rol = self._v_rol.get()
            cfg = ROLES.get(rol, ROLES["Constructor"])
            self._prev_lbl.config(
                text=f"{cfg['emoji']}  {rol}  —  {cfg['desc']}",
                fg=cfg["color"])
            if PIL_OK:
                nombre = self._v_nombre.get() or "?"
                self._avatar_prev = AvatarGen.generar(rol, nombre, 60)
                if self._avatar_prev:
                    self._prev_avatar_lbl.config(image=self._avatar_prev)

        self._v_rol.trace_add("write", update_preview)
        self._v_nombre.trace_add("write", update_preview)
        update_preview()

        sep(form)

        btn_f = tk.Frame(form, bg=BG)
        btn_f.pack(pady=8)
        mk_btn(btn_f, "💾 Guardar", self._guardar, GOLD, BG).pack(side=tk.LEFT, padx=6)
        mk_btn(btn_f, "Cancelar",   self.destroy, BG3, GRAY).pack(side=tk.LEFT, padx=6)

    def _guardar(self):
        nombre = self._v_nombre.get().strip()
        if not nombre:
            messagebox.showerror("Error", "El nombre es obligatorio.", parent=self)
            return
        try:
            exp = int(self._v_exp.get())
        except ValueError:
            messagebox.showerror("Error", "EXP debe ser un número entero.", parent=self)
            return

        nuevo = {
            "id":           self.datos["id"] if self.datos else self.db.siguiente_id(),
            "nombre":       nombre,
            "rol":          self._v_rol.get(),
            "exp":          exp,
            "kinket":       self.datos.get("kinket", []) if self.datos else [],
            "quests":       self.datos.get("quests", []) if self.datos else [],
            "fecha_ingreso":self.datos.get("fecha_ingreso", str(datetime.date.today())) if self.datos else str(datetime.date.today()),
            "notas":        self._v_notas.get().strip(),
        }
        self.db.guardar(nuevo)
        if self.on_save: self.on_save()
        self.destroy()


# ══════════════════════════════════════════════════════════════════
#  APP PRINCIPAL
# ══════════════════════════════════════════════════════════════════

class BuilderGuildApp(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title("🏰  Builder Guild  —  Panel de Constructores")
        self.configure(bg=BG)
        self.geometry("1060x720")
        self.minsize(900, 600)

        self.db           = ExcelDB(EXCEL_PATH)
        self._avatares    = {}
        self._busqueda    = tk.StringVar()
        self._filtro_rol  = tk.StringVar(value="Todos")
        self._busqueda.trace_add("write",   self._on_filtro)
        self._filtro_rol.trace_add("write", self._on_filtro)

        self._build_ui()
        self._cargar_lista()

    # ── BUILD UI ──────────────────────────────────────────────────
    def _build_ui(self):
        # TOP BAR
        top = tk.Frame(self, bg=BG2, height=56)
        top.pack(fill=tk.X)
        top.pack_propagate(False)

        tk.Label(top, text="🏰  BUILDER GUILD",
                 bg=BG2, fg=GOLD,
                 font=("Georgia", 18, "bold")).pack(side=tk.LEFT, padx=18, pady=12)
        tk.Label(top, text="Panel de Constructores",
                 bg=BG2, fg=GRAY,
                 font=("Helvetica", 9, "italic")).pack(side=tk.LEFT, padx=0, pady=16)

        # Botón abrir Excel
        mk_btn(top, "📊 Abrir Excel", self._abrir_excel,
               BG3, GRAY, font=("Helvetica",8,"bold")).pack(side=tk.RIGHT, padx=8, pady=12)
        mk_btn(top, "➕ Nuevo Constructor", self._nuevo_constructor,
               GREEN, BG, font=("Helvetica",9,"bold")).pack(side=tk.RIGHT, padx=4, pady=12)

        tk.Frame(self, bg=BORDER, height=1).pack(fill=tk.X)

        # FILTROS
        filt = tk.Frame(self, bg=BG3, height=44)
        filt.pack(fill=tk.X)
        filt.pack_propagate(False)

        tk.Label(filt, text="🔍", bg=BG3, fg=GRAY,
                 font=("Helvetica",12)).pack(side=tk.LEFT, padx=(14,4), pady=8)
        entry = tk.Entry(filt, textvariable=self._busqueda,
                         bg=BG2, fg=WHITE, insertbackground=GOLD,
                         relief="solid", font=("Helvetica",10),
                         bd=1, width=22)
        entry.pack(side=tk.LEFT, ipady=4, pady=8)

        tk.Label(filt, text="Rol:", bg=BG3, fg=GRAY,
                 font=("Helvetica",9)).pack(side=tk.LEFT, padx=(14,4))
        roles_opts = ["Todos"] + list(ROLES.keys())
        rol_cb = ttk.Combobox(filt, textvariable=self._filtro_rol,
                               values=roles_opts, state="readonly",
                               width=16, font=("Helvetica",9))
        rol_cb.pack(side=tk.LEFT, pady=8)

        # Stats rápidos
        self._stat_total_lbl = tk.Label(filt, text="", bg=BG3,
                                         fg=GRAY, font=("Helvetica",9))
        self._stat_total_lbl.pack(side=tk.RIGHT, padx=16)

        tk.Frame(self, bg=BORDER, height=1).pack(fill=tk.X)

        # LISTA + DETALLE
        main = tk.Frame(self, bg=BG)
        main.pack(fill=tk.BOTH, expand=True)

        # Panel lista (izquierda)
        list_panel = tk.Frame(main, bg=BG, width=340)
        list_panel.pack(side=tk.LEFT, fill=tk.Y)
        list_panel.pack_propagate(False)

        # Canvas scrollable para las cards
        self._list_canvas = tk.Canvas(list_panel, bg=BG,
                                       highlightthickness=0, width=336)
        vsb = ttk.Scrollbar(list_panel, orient="vertical",
                             command=self._list_canvas.yview)
        self._list_canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self._list_canvas.pack(fill=tk.BOTH, expand=True)

        self._list_inner = tk.Frame(self._list_canvas, bg=BG)
        self._list_win   = self._list_canvas.create_window(
            (0,0), window=self._list_inner, anchor="nw")
        self._list_inner.bind("<Configure>",
            lambda e: self._list_canvas.configure(
                scrollregion=self._list_canvas.bbox("all")))
        self._list_canvas.bind("<Configure>",
            lambda e: self._list_canvas.itemconfig(self._list_win, width=e.width))
        self._list_canvas.bind("<MouseWheel>",
            lambda e: self._list_canvas.yview_scroll(-1*(e.delta//120), "units"))

        # Separador
        tk.Frame(main, bg=BORDER, width=1).pack(side=tk.LEFT, fill=tk.Y)

        # Panel derecho: ranking + resumen
        self._right_panel = tk.Frame(main, bg=BG)
        self._right_panel.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self._build_right_panel()

    def _build_right_panel(self):
        rp = self._right_panel
        for w in rp.winfo_children(): w.destroy()

        todos = self.db.leer_todos()

        # Header
        tk.Label(rp, text="🏆  RANKING DEL GREMIO",
                 bg=BG, fg=GOLD,
                 font=("Georgia",14,"bold")).pack(pady=(16,4), padx=16, anchor="w")
        tk.Frame(rp, bg=GOLD, height=1).pack(fill=tk.X, padx=16, pady=(0,8))

        if not todos:
            tk.Label(rp, text="Sin constructores. ¡Agrega el primero!",
                     bg=BG, fg=GRAY,
                     font=("Helvetica",11,"italic")).pack(pady=40)
            return

        # Top 3 podio
        top3 = sorted(todos, key=lambda x: x["exp"], reverse=True)[:3]
        podio_f = tk.Frame(rp, bg=BG)
        podio_f.pack(fill=tk.X, padx=16, pady=(0,16))

        medallas = ["🥇", "🥈", "🥉"]
        for i, c in enumerate(top3):
            rol = c["rol"]
            cfg = ROLES.get(rol, ROLES["Constructor"])
            nv_num, nv_nom, *_ = get_nivel(c["exp"])

            card = tk.Frame(podio_f, bg=BG2, padx=10, pady=10,
                            highlightthickness=1,
                            highlightbackground=cfg["color"])
            card.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=4)

            tk.Label(card, text=medallas[i], bg=BG2,
                     font=("Segoe UI Emoji", 20)).pack()

            if PIL_OK:
                av = AvatarGen.generar_mini(rol, c["nombre"], 50)
                if av:
                    self._avatares[f"top_{c['id']}"] = av
                    tk.Label(card, image=av, bg=BG2).pack()

            tk.Label(card, text=c["nombre"][:14],
                     bg=BG2, fg=cfg["color"],
                     font=("Helvetica",9,"bold")).pack()
            tk.Label(card, text=f"{cfg['emoji']} {rol}",
                     bg=BG2, fg=GRAY,
                     font=("Helvetica",7)).pack()
            tk.Label(card, text=f"⭐ {c['exp']:,} XP",
                     bg=BG2, fg=GOLD,
                     font=("Courier New",10,"bold")).pack()
            tk.Label(card, text=f"Nv.{nv_num} {nv_nom}",
                     bg=BG2, fg=GRAY,
                     font=("Helvetica",7)).pack()

        # Tabla ranking completo
        tk.Label(rp, text="Todos los constructores:",
                 bg=BG, fg=GRAY,
                 font=("Helvetica",9)).pack(padx=16, anchor="w")

        tabla_wrap = tk.Frame(rp, bg=BG)
        tabla_wrap.pack(fill=tk.BOTH, expand=True, padx=16, pady=(4,16))

        # Treeview
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Guild.Treeview",
                        background=BG2, fieldbackground=BG2,
                        foreground=WHITE, rowheight=28,
                        font=("Helvetica",9))
        style.configure("Guild.Treeview.Heading",
                        background=BG3, foreground=GOLD,
                        font=("Helvetica",9,"bold"))
        style.map("Guild.Treeview",
                  background=[("selected", BG3)],
                  foreground=[("selected", GOLD)])

        cols = ("#", "Nombre", "Rol", "EXP", "Nivel", "Quests", "Kinket")
        tree = ttk.Treeview(tabla_wrap, columns=cols,
                            show="headings", style="Guild.Treeview")

        wsb = ttk.Scrollbar(tabla_wrap, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=wsb.set)
        wsb.pack(side=tk.RIGHT, fill=tk.Y)
        tree.pack(fill=tk.BOTH, expand=True)

        widths = [32, 160, 110, 80, 90, 60, 60]
        for col, w in zip(cols, widths):
            tree.heading(col, text=col)
            tree.column(col, width=w, anchor="center" if col!="Nombre" else "w")

        todos_sorted = sorted(todos, key=lambda x: x["exp"], reverse=True)
        for i, c in enumerate(todos_sorted, 1):
            _, nv_nom, nv_emoji, *_ = get_nivel(c["exp"])
            cfg = ROLES.get(c["rol"], ROLES["Constructor"])
            tree.insert("", "end", iid=str(c["id"]), values=(
                str(i),
                f"  {c['nombre']}",
                f"{cfg['emoji']} {c['rol']}",
                f"{c['exp']:,}",
                f"{nv_emoji} {nv_nom}",
                str(len(c["quests"])),
                str(len(c["kinket"])),
            ))

        tree.bind("<Double-1>",
                  lambda e: self._abrir_detalle_desde_tree(tree))

    def _abrir_detalle_desde_tree(self, tree):
        sel = tree.selection()
        if sel:
            DetalleWindow(self, self.db, int(sel[0]),
                          on_close=self._refrescar)

    # ── LISTA DE CARDS ────────────────────────────────────────────
    def _cargar_lista(self):
        self._todos = self.db.leer_todos()
        self._on_filtro()

    def _on_filtro(self, *_):
        q    = self._busqueda.get().lower().strip()
        rol  = self._filtro_rol.get()
        filtrados = [
            c for c in getattr(self, "_todos", [])
            if (q in c["nombre"].lower() or q in c["rol"].lower() or not q)
            and (rol == "Todos" or c["rol"] == rol)
        ]
        self._render_cards(filtrados)
        total_exp = sum(c["exp"] for c in filtrados)
        self._stat_total_lbl.config(
            text=f"{len(filtrados)} constructores  |  {total_exp:,} XP total")

    def _render_cards(self, constructores: list[dict]):
        for w in self._list_inner.winfo_children():
            w.destroy()
        self._avatares_cards = {}

        for c in constructores:
            self._render_card(c)

    def _render_card(self, c: dict):
        rol    = c["rol"]
        cfg    = ROLES.get(rol, ROLES["Constructor"])
        col    = cfg["color"]
        nv_num, nv_nom, nv_emoji, exp_en_nv, exp_para_nv = get_nivel(c["exp"])
        pct    = exp_en_nv / exp_para_nv if exp_para_nv > 0 else 1.0

        card = tk.Frame(self._list_inner, bg=BG2, padx=10, pady=8,
                        cursor="hand2",
                        highlightthickness=1,
                        highlightbackground=BORDER)
        card.pack(fill=tk.X, padx=6, pady=3)

        def hover_in(e, f=card, c2=col):
            f.config(bg=BG3, highlightbackground=c2)
            for w in f.winfo_children(): _update_bg(w, BG3)
        def hover_out(e, f=card):
            f.config(bg=BG2, highlightbackground=BORDER)
            for w in f.winfo_children(): _update_bg(w, BG2)
        def _update_bg(w, bg):
            try: w.config(bg=bg)
            except: pass
            for ch in w.winfo_children():
                _update_bg(ch, bg)

        card.bind("<Enter>", hover_in)
        card.bind("<Leave>", hover_out)
        card.bind("<Button-1>", lambda e, cid=c["id"]: self._abrir_detalle(cid))

        # Row principal
        row = tk.Frame(card, bg=BG2)
        row.pack(fill=tk.X)

        # Avatar mini
        if PIL_OK:
            av = AvatarGen.generar_mini(rol, c["nombre"], 44)
            if av:
                self._avatares_cards[c["id"]] = av
                av_lbl = tk.Label(row, image=av, bg=BG2, cursor="hand2")
                av_lbl.pack(side=tk.LEFT, padx=(0,8))
                av_lbl.bind("<Button-1>", lambda e, cid=c["id"]: self._abrir_detalle(cid))

        # Info
        info = tk.Frame(row, bg=BG2)
        info.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        info.bind("<Button-1>", lambda e, cid=c["id"]: self._abrir_detalle(cid))

        name_row = tk.Frame(info, bg=BG2)
        name_row.pack(anchor="w")
        tk.Label(name_row, text=c["nombre"], bg=BG2, fg=WHITE,
                 font=("Helvetica",11,"bold"),
                 cursor="hand2").pack(side=tk.LEFT)
        tk.Label(name_row,
                 text=f"  {nv_emoji} Nv.{nv_num}",
                 bg=BG2, fg=col,
                 font=("Helvetica",8)).pack(side=tk.LEFT)

        tk.Label(info, text=f"{cfg['emoji']}  {rol}",
                 bg=BG2, fg=col,
                 font=("Helvetica",8)).pack(anchor="w")

        # Chips de stats
        chips = tk.Frame(info, bg=BG2)
        chips.pack(anchor="w", pady=(2,0))
        for txt, c_txt in [
            (f"⭐ {c['exp']:,} XP", GOLD),
            (f"🎯 {len(c['quests'])}", CYAN),
            (f"🎒 {len(c['kinket'])}", PURPLE),
        ]:
            tk.Label(chips, text=txt, bg=DIM, fg=c_txt,
                     font=("Helvetica",7,"bold"),
                     padx=5, pady=1).pack(side=tk.LEFT, padx=2)

        # Mini barra XP
        bar_row = tk.Frame(card, bg=BG2)
        bar_row.pack(fill=tk.X, pady=(4,0))
        mini_bar = XPBar(bar_row, width=280, height=5)
        mini_bar.pack(side=tk.LEFT, padx=2)
        mini_bar.set_value(pct, col, animate=False)

        # Bind para todos los hijos
        for w in [row, info, bar_row, chips, name_row]:
            w.bind("<Enter>", hover_in)
            w.bind("<Leave>", hover_out)
            w.bind("<Button-1>", lambda e, cid=c["id"]: self._abrir_detalle(cid))

    def _abrir_detalle(self, cid: int):
        DetalleWindow(self, self.db, cid,
                      on_close=self._refrescar)

    def _refrescar(self):
        self._todos = self.db.leer_todos()
        self._on_filtro()
        self._build_right_panel()

    def _nuevo_constructor(self):
        EditarWindow(self, self.db, None,
                     on_save=self._refrescar)

    def _abrir_excel(self):
        import subprocess, sys
        try:
            if sys.platform == "win32":
                os.startfile(str(EXCEL_PATH))
            elif sys.platform == "darwin":
                subprocess.Popen(["open", str(EXCEL_PATH)])
            else:
                subprocess.Popen(["xdg-open", str(EXCEL_PATH)])
        except Exception:
            messagebox.showinfo("Ruta del archivo",
                                f"El archivo Excel está en:\n{EXCEL_PATH}")


# ══════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    try:
        import openpyxl
    except ImportError:
        print("Instala openpyxl:  pip install openpyxl")
        exit(1)

    app = BuilderGuildApp()
    app.mainloop()
